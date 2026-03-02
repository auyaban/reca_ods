import json

import os
import time

import shutil

from datetime import datetime
from uuid import uuid4

from typing import Any



from copy import copy



from app.logging_utils import LOGGER_EXCEL, get_file_logger
from app.paths import app_data_dir, resource_path

from app.storage import ensure_appdata_files, _desktop_excel_dir
from app.utils.text import normalize_text

try:
    from postgrest.exceptions import APIError as PostgrestAPIError
except ImportError:  # pragma: no cover - dependencia opcional en runtime
    PostgrestAPIError = RuntimeError  # type: ignore[assignment]

from app.factura_calc import calcular_items
from app.constants import (
    FACTURA_GRAND_TOTAL_CELL,
    FACTURA_HEADER_ROW,
    FACTURA_IVA_CELL,
    FACTURA_TOTAL_CELL,
    IVA_RATE,
)



_DATA_ROOT = _desktop_excel_dir()

_EXCEL_PATH = _DATA_ROOT / "ODS 2026.xlsx"

_EXCEL_QUEUE = _DATA_ROOT / "ODS 2026 pendiente.jsonl"

_ODS_SHEET_GENERAL = "ODS General"
_ODS_SHEET_FILTRADA = "ODS Filtrada"
_YEAR_FIELD = "ano_servicio"
_LEGACY_YEAR_FIELDS = ("a\u00f1o_servicio", "a\u00c3\u00b1o_servicio")



_LOG_FILE = app_data_dir() / "logs" / "excel.log"
_logger = get_file_logger(LOGGER_EXCEL, _LOG_FILE, announce=True)
_SUPABASE_FETCH_ERRORS = (
    PostgrestAPIError,
    RuntimeError,
    ValueError,
    TypeError,
    OSError,
)


def _new_op_id() -> str:
    return uuid4().hex[:8]


def _log_event(level: str, context: str, op_id: str, message: str, *args, **kwargs) -> None:
    log_fn = getattr(_logger, level, _logger.info)
    log_fn(f"[ctx={context} op={op_id}] {message}", *args, **kwargs)





def _normalize_header(value: str) -> str:
    return normalize_text(value, lowercase=True)





_EXCEL_FIELD_MAP = {

    "id": "id",

    "profesional": "nombre_profesional",

    "nuevo codigo": "codigo_servicio",

    "empresa": "nombre_empresa",

    "nit": "nit_empresa",

    "ccf": "caja_compensacion",

    "fecha": "fecha_servicio",

    "referencia": "referencia_servicio",

    "nombre": "descripcion_servicio",

    "oferentes": "nombre_usuario",

    "cedula": "cedula_usuario",

    "tipo de discapacidad": "discapacidad_usuario",

    "fecha ingreso": "fecha_ingreso",

    "valor servicio virtual": "valor_virtual",

    "valor servicio bogota": "valor_bogota",

    "valor fuera de bogota": "valor_otro",

    "todas las modalidades": "todas_modalidades",

    "total horas": "horas_interprete",

    "valor a pagar": "valor_interprete",

    "total valor servicio sin iva": "valor_total",

    "observaciones": "observaciones",

    "asesor": "asesor_empresa",

    "sede": "sede_empresa",

    "modalidad": "modalidad_servicio",

    "observacion agencia": "observacion_agencia",

    "clausulada": "orden_clausulada",

    "a\u00f1o": _YEAR_FIELD,

    "ano": _YEAR_FIELD,

    "mes": "mes_servicio",

    "genero": "genero_usuario",

    "tipo de contrato": "tipo_contrato",

    "seguimiento": "seguimiento_servicio",

    "cargo": "cargo_servicio",

    "personas": "total_personas",

}



_MATCH_KEYS = [

    "id",

    "fecha_servicio",

    "codigo_servicio",

    "nit_empresa",

    "nombre_profesional",

]







def _get_year_value(data: dict) -> int:
    for key in (_YEAR_FIELD, *_LEGACY_YEAR_FIELDS):

        value = data.get(key)

        if value not in (None, ""):

            return int(value)

    return 0



def _coerce_to_string(value: Any) -> str:

    if value is None:

        return ""

    return str(value).strip()





def _row_to_ods(ws, row_idx: int, headers: list[str], normalized_headers: list[str]) -> dict:

    row_data = {}

    for col_idx, header_key in enumerate(normalized_headers, start=1):

        field = _EXCEL_FIELD_MAP.get(header_key)

        if not field:

            continue

        value = ws.cell(row=row_idx, column=col_idx).value

        row_data[field] = value

    return row_data





def _match_row(original: dict, row_data: dict) -> bool:

    for key in _MATCH_KEYS:

        if _coerce_to_string(original.get(key)) != _coerce_to_string(row_data.get(key)):

            return False

    return True





def _find_target_row(ws, original: dict, headers: list[str], normalized_headers: list[str]) -> int | None:

    for row_idx in range(2, ws.max_row + 1):

        if all(cell.value in (None, "") for cell in ws[row_idx]):

            continue

        row_data = _row_to_ods(ws, row_idx, headers, normalized_headers)

        if _match_row(original, row_data):

            return row_idx

    return None






def _prepare_sheet(ws: Any, get_column_letter: Any) -> tuple[list[str], list[str]]:
    headers = [str(cell.value or "") for cell in ws[1]]
    normalized_headers = [_normalize_header(str(h)) for h in headers]
    if "id" not in normalized_headers:
        new_col = len(headers) + 1
        ws.cell(row=1, column=new_col, value="id")
        headers.append("id")
        normalized_headers.append("id")
        ws.column_dimensions[get_column_letter(new_col)].hidden = True
    return headers, normalized_headers


def _get_target_sheets(wb: Any, get_column_letter: Any) -> list[tuple[Any, list[str], list[str]]]:
    targets = []
    lower_map = {name.lower(): name for name in wb.sheetnames}
    for name in (_ODS_SHEET_GENERAL, _ODS_SHEET_FILTRADA):
        actual_name = lower_map.get(name.lower())
        if actual_name:
            ws = wb[actual_name]
            headers, normalized_headers = _prepare_sheet(ws, get_column_letter)
            targets.append((ws, headers, normalized_headers))
    if not targets:
        ws = wb.active
        headers, normalized_headers = _prepare_sheet(ws, get_column_letter)
        targets.append((ws, headers, normalized_headers))
    return targets


def _load_excel(op_id: str | None = None) -> tuple[Any, list[tuple[Any, list[str], list[str]]]]:
    op_id = op_id or _new_op_id()
    context = "excel.load"

    ensure_appdata_files()

    try:

        import openpyxl

        from openpyxl.utils import get_column_letter

    except ImportError as exc:  # pragma: no cover

        raise RuntimeError("openpyxl no esta instalado") from exc



    if not _EXCEL_PATH.exists():
        _log_event("error", context, op_id, "Excel no encontrado. Ruta=%s", _EXCEL_PATH)

        raise RuntimeError(f"No se encontro el archivo Excel: {_EXCEL_PATH}")

    _log_event("info", context, op_id, "Abriendo Excel. Ruta=%s", _EXCEL_PATH)

    wb = openpyxl.load_workbook(_EXCEL_PATH)

    targets = _get_target_sheets(wb, get_column_letter)
    return wb, targets


def _safe_save_workbook(wb, op_id: str | None = None) -> None:
    op_id = op_id or _new_op_id()
    context = "excel.save"

    tmp_path = _EXCEL_PATH.with_suffix(".tmp")

    if tmp_path.exists():

        tmp_path.unlink()

    try:
        wb.save(tmp_path)
    except (OSError, RuntimeError, ValueError) as exc:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError as unlink_exc:
                _log_event("warning", context, op_id, "No se pudo limpiar temporal tras fallo de save: %s", unlink_exc)
        try:
            wb.close()
        except (OSError, RuntimeError, ValueError) as close_exc:
            _log_event("exception", context, op_id, "Fallo al cerrar workbook tras error de save temporal.")
            raise RuntimeError(
                f"No se pudo escribir archivo temporal de Excel ({exc}) y tampoco cerrar workbook ({close_exc})"
            ) from close_exc
        _log_event("exception", context, op_id, "Fallo al guardar archivo temporal de Excel. Ruta=%s", tmp_path)
        raise RuntimeError(f"No se pudo escribir archivo temporal de Excel: {exc}") from exc

    try:
        wb.close()
    except (OSError, RuntimeError, ValueError) as exc:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError as unlink_exc:
                _log_event("warning", context, op_id, "No se pudo limpiar temporal tras fallo de close: %s", unlink_exc)
        _log_event("exception", context, op_id, "Fallo al cerrar workbook antes de replace.")
        raise RuntimeError(f"No se pudo cerrar workbook antes de guardar Excel: {exc}") from exc

    last_error = None
    for attempt in range(1, 5):
        try:
            os.replace(tmp_path, _EXCEL_PATH)
            _log_event("info", context, op_id, "Excel guardado. Ruta=%s Intento=%s", _EXCEL_PATH, attempt)
            return
        except PermissionError as exc:
            last_error = exc
            if attempt >= 4:
                break
            _log_event(
                "warning",
                context,
                op_id,
                "Excel bloqueado durante replace. Reintentando... intento=%s ruta=%s",
                attempt,
                _EXCEL_PATH,
            )
            time.sleep(0.2 * attempt)
        except OSError as exc:
            last_error = exc
            if attempt >= 4:
                break
            # WinError 32/33: archivo en uso o bloqueo compartido.
            if getattr(exc, "winerror", None) in {32, 33}:
                _log_event(
                    "warning",
                    context,
                    op_id,
                    "Excel en uso durante replace (winerror=%s). Reintentando... intento=%s",
                    getattr(exc, "winerror", None),
                    attempt,
                )
                time.sleep(0.2 * attempt)
                continue
            break
        except (RuntimeError, ValueError) as exc:
            last_error = exc
            break

    if tmp_path.exists():
        try:
            tmp_path.unlink()
        except OSError as exc:
            _log_event("warning", context, op_id, "No se pudo eliminar temporal tras fallo de replace: %s", exc)

    _log_event("error", context, op_id, "Fallo al guardar Excel. Ruta=%s", _EXCEL_PATH, exc_info=last_error)
    if isinstance(last_error, (PermissionError, OSError)):
        raise PermissionError(f"No se pudo guardar Excel porque el archivo esta en uso: {_EXCEL_PATH}") from last_error
    raise RuntimeError(f"No se pudo guardar Excel: {last_error}") from last_error





def _nombre_hoja_factura(mes: int, ano: int, tipo: str) -> str:

    meses = [

        "Ene",

        "Feb",

        "Mar",

        "Abr",

        "May",

        "Jun",

        "Jul",

        "Ago",

        "Sep",

        "Oct",

        "Nov",

        "Dic",

    ]

    nombre_mes = meses[mes - 1] if 1 <= mes <= 12 else "Mes"

    tipo_label = "Claus" if tipo.strip().lower() == "clausulada" else "NoClaus"

    nombre = f"Factura {nombre_mes} {ano} {tipo_label}"

    return nombre[:31]





def _render_factura_sheet(wb: Any, mes: int, ano: int, tipo: str) -> None:
    op_id = _new_op_id()
    context = "factura.render"

    try:

        import openpyxl

    except ImportError as exc:  # pragma: no cover

        raise RuntimeError("openpyxl no esta instalado") from exc



    template_name = "clausulada.xlsx" if tipo.strip().lower() == "clausulada" else "no_clausulada.xlsx"

    template_path = resource_path(f"facturas/{template_name}")

    _log_event(
        "info",
        context,
        op_id,

        "Actualizando factura. Mes=%s Ano=%s Tipo=%s Plantilla=%s",

        mes,

        ano,

        tipo,

        template_path,

    )

    template_wb = openpyxl.load_workbook(template_path)
    try:
        template_ws = template_wb.active
        items = calcular_items(mes, ano, tipo)

        sheet_name = _nombre_hoja_factura(mes, ano, tipo)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        for row in template_ws.iter_rows():
            for cell in row:
                if cell.__class__.__name__ == "MergedCell":
                    continue

                target = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    target._style = copy(cell._style)
                target.number_format = cell.number_format
                target.alignment = copy(cell.alignment)
                target.border = copy(cell.border)
                target.fill = copy(cell.fill)
                target.font = copy(cell.font)

        for merged in template_ws.merged_cells.ranges:
            ws.merge_cells(str(merged))

        for key, dim in template_ws.row_dimensions.items():
            ws.row_dimensions[key].height = dim.height

        for key, dim in template_ws.column_dimensions.items():
            ws.column_dimensions[key].width = dim.width

        header_row = FACTURA_HEADER_ROW
        for idx, item in enumerate(items, start=1):
            row_idx = header_row + (idx - 1)
            ws.cell(row=row_idx, column=1, value=item.codigo_servicio)
            ws.cell(row=row_idx, column=2, value=item.referencia_servicio)
            ws.cell(row=row_idx, column=3, value=item.descripcion_servicio)
            ws.cell(row=row_idx, column=4, value=item.valor_base)
            ws.cell(row=row_idx, column=5, value=item.cantidad)
            ws.cell(row=row_idx, column=6, value=item.total)

        total_sum = sum(item.total for item in items)
        ws[FACTURA_TOTAL_CELL] = total_sum
        ws[FACTURA_IVA_CELL] = round(total_sum * IVA_RATE, 2)
        ws[FACTURA_GRAND_TOTAL_CELL] = round(total_sum + ws[FACTURA_IVA_CELL].value, 2)
    finally:
        template_wb.close()





def update_factura_sheet(mes: int, ano: int, tipo: str) -> None:
    op_id = _new_op_id()
    context = "factura.update"

    ensure_appdata_files()

    try:

        import openpyxl

    except ImportError as exc:  # pragma: no cover

        raise RuntimeError("openpyxl no esta instalado") from exc



    wb = openpyxl.load_workbook(_EXCEL_PATH)
    saved = False
    try:
        _render_factura_sheet(wb, mes, ano, tipo)
        _safe_save_workbook(wb, op_id=op_id)
        saved = True
    finally:
        if not saved:
            try:
                wb.close()
            except (OSError, RuntimeError, ValueError) as exc:
                _log_event("warning", context, op_id, "No se pudo cerrar workbook tras fallo en update_factura_sheet: %s", exc)







def _build_row_values(ods_data: dict, headers: list[str], normalized_headers: list[str]) -> list[Any]:

    row_values = [None] * len(headers)

    for idx, header_key in enumerate(normalized_headers):

        if header_key == "#":

            row_values[idx] = None

            continue

        field = _EXCEL_FIELD_MAP.get(header_key)

        if not field:

            continue

        value = ods_data.get(field)

        if value is None and field == _YEAR_FIELD:
            for legacy_key in _LEGACY_YEAR_FIELDS:
                value = ods_data.get(legacy_key)
                if value is not None:
                    break

        if field == "orden_clausulada":
            value = "Si" if str(value).strip().lower().startswith("s") else "No"

        row_values[idx] = value

    return row_values





def append_row(ods_data: dict) -> None:
    op_id = _new_op_id()
    context = "excel.append"
    wb, targets = _load_excel(op_id=op_id)
    saved = False
    try:
        for ws, headers, normalized_headers in targets:
            row_values = _build_row_values(ods_data, headers, normalized_headers)
            target_row = None
            for row_idx in range(2, ws.max_row + 1):
                cells = ws[row_idx]
                if all(cell.value in (None, "") for cell in cells):
                    target_row = row_idx
                    break
            if target_row is None:
                target_row = ws.max_row + 1
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)
        _safe_save_workbook(wb, op_id=op_id)
        saved = True
    finally:
        if not saved:
            try:
                wb.close()
            except (OSError, RuntimeError, ValueError) as exc:
                _log_event("warning", context, op_id, "No se pudo cerrar workbook tras fallo en append_row: %s", exc)





def update_row(original: dict, ods_data: dict) -> None:
    op_id = _new_op_id()
    context = "excel.update"
    wb, targets = _load_excel(op_id=op_id)
    saved = False
    try:
        found_any = False
        target_rows: list[tuple[Any, int | None, list[Any]]] = []
        for ws, headers, normalized_headers in targets:
            target_row = _find_target_row(ws, original, headers, normalized_headers)
            row_values = _build_row_values(ods_data, headers, normalized_headers)
            if target_row is not None:
                found_any = True
            target_rows.append((ws, target_row, row_values))

        if not found_any:
            raise RuntimeError("No se encontro la fila en Excel para actualizar")

        for ws, target_row, row_values in target_rows:
            if target_row is None:
                _log_event("warning", context, op_id, "Fila no encontrada en hoja %s; se agregara para mantener sincronia.", ws.title)
                for row_idx in range(2, ws.max_row + 1):
                    cells = ws[row_idx]
                    if all(cell.value in (None, "") for cell in cells):
                        target_row = row_idx
                        break
                if target_row is None:
                    target_row = ws.max_row + 1

            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)

        _safe_save_workbook(wb, op_id=op_id)
        saved = True
    finally:
        if not saved:
            try:
                wb.close()
            except (OSError, RuntimeError, ValueError) as exc:
                _log_event("warning", context, op_id, "No se pudo cerrar workbook tras fallo en update_row: %s", exc)





def delete_row(original: dict) -> None:
    op_id = _new_op_id()
    context = "excel.delete"
    wb, targets = _load_excel(op_id=op_id)
    saved = False
    try:
        found_any = False
        for ws, headers, normalized_headers in targets:
            target_row = _find_target_row(ws, original, headers, normalized_headers)
            if target_row is None:
                continue
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=target_row, column=col_idx, value=None)
            found_any = True
        if not found_any:
            raise RuntimeError("No se encontro la fila en Excel para eliminar")
        _safe_save_workbook(wb, op_id=op_id)
        saved = True
    finally:
        if not saved:
            try:
                wb.close()
            except (OSError, RuntimeError, ValueError) as exc:
                _log_event("warning", context, op_id, "No se pudo cerrar workbook tras fallo en delete_row: %s", exc)







def rebuild_excel_from_supabase(rows: list[dict], create_backup: bool = True) -> dict:
    op_id = _new_op_id()
    context = "excel.rebuild"

    ensure_appdata_files()

    try:

        import openpyxl

        from openpyxl.utils import get_column_letter

    except ImportError as exc:  # pragma: no cover

        raise RuntimeError("openpyxl no esta instalado") from exc



    _log_event("info", context, op_id, "Rebuild Excel iniciado. Filas=%s", len(rows))



    backup_path = None

    if create_backup and _EXCEL_PATH.exists():

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        backup_path = _EXCEL_PATH.with_name(f"ODS 2026 backup {timestamp}.xlsx")

        shutil.copy2(_EXCEL_PATH, backup_path)



    template_path = resource_path("Excel/ods_2026.xlsx")

    wb = openpyxl.load_workbook(template_path)
    saved = False
    try:
        targets = _get_target_sheets(wb, get_column_letter)
        for ws, headers, normalized_headers in targets:
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx, value=None)

            target_row = 2
            for row in rows:
                row_values = _build_row_values(row, headers, normalized_headers)
                for col_idx, value in enumerate(row_values, start=1):
                    ws.cell(row=target_row, column=col_idx, value=value)
                target_row += 1

        _safe_save_workbook(wb, op_id=op_id)
        saved = True
    finally:
        if not saved:
            try:
                wb.close()
            except (OSError, RuntimeError, ValueError) as exc:
                _log_event("warning", context, op_id, "No se pudo cerrar workbook tras fallo en rebuild_excel_from_supabase: %s", exc)

    _log_event("info", context, op_id, "Rebuild Excel finalizado. Filas=%s", len(rows))

    return {"rows": len(rows), "backup": str(backup_path) if backup_path else ""}





def _fetch_all_ods_rows(page_size: int = 1000) -> list[dict]:
    from app.supabase_client import get_supabase_client

    op_id = _new_op_id()
    context = "excel.rebuild.fetch"
    safe_page_size = max(1, int(page_size))
    client = get_supabase_client()
    rows: list[dict] = []
    last_id: int | None = None

    try:
        while True:
            query = client.table("ods").select("*").order("id").limit(safe_page_size)
            if last_id is not None:
                query = query.gt("id", last_id)
            response = query.execute()
            batch = list(response.data or [])
            if not batch:
                break
            rows.extend(batch)
            last_raw = batch[-1].get("id")
            if last_raw in (None, ""):
                raise RuntimeError("No se pudo paginar ODS por keyset: columna id ausente en respuesta.")
            next_last_id = int(last_raw)
            if last_id is not None and next_last_id <= last_id:
                raise RuntimeError("No se pudo paginar ODS por keyset: id no monotono en respuesta.")
            last_id = next_last_id
            if len(batch) < safe_page_size:
                break
    except _SUPABASE_FETCH_ERRORS as exc:
        _log_event(
            "exception",
            context,
            op_id,
            "Fallo leyendo ODS desde Supabase. LastId=%s PageSize=%s",
            last_id,
            safe_page_size,
        )
        raise RuntimeError(f"No se pudieron obtener filas de Supabase: {exc}") from exc

    _log_event(
        "info",
        context,
        op_id,
        "Filas ODS descargadas para rebuild: %s",
        len(rows),
    )
    return rows


def rebuild_excel_from_supabase_query(create_backup: bool = True) -> None:
    op_id = _new_op_id()
    context = "excel.rebuild.query"
    rows = _fetch_all_ods_rows(page_size=1000)
    _log_event(
        "info",
        context,
        op_id,
        "Iniciando rebuild desde query paginada. Filas=%s",
        len(rows),
    )
    rebuild_excel_from_supabase(rows, create_backup=create_backup)





def queue_action(action: str, ods: dict, original: dict | None, reason: str, meta: dict | None = None) -> None:
    op_id = _new_op_id()
    context = "excel.queue"

    _EXCEL_QUEUE.parent.mkdir(parents=True, exist_ok=True)

    record = {

        "timestamp": datetime.utcnow().isoformat(),

        "action": action,

        "reason": reason,

        "ods": ods,

        "original": original,

        "meta": meta or {},

    }

    with _EXCEL_QUEUE.open("a", encoding="utf-8") as handle:

        handle.write(json.dumps(record, ensure_ascii=True) + "\n")

    _log_event("info", context, op_id, "Encolado Excel. Accion=%s Motivo=%s Ruta=%s", action, reason, _EXCEL_QUEUE)





def clear_queue() -> None:

    if _EXCEL_QUEUE.exists():

        _EXCEL_QUEUE.write_text("", encoding="utf-8")





def get_queue_status() -> dict:

    pendientes = 0

    if _EXCEL_QUEUE.exists():

        lines = _EXCEL_QUEUE.read_text(encoding="utf-8").splitlines()

        pendientes = len([line for line in lines if line.strip()])

    return {"pendientes": pendientes}





def flush_queue() -> dict:
    op_id = _new_op_id()
    context = "excel.flush"

    if not _EXCEL_QUEUE.exists():
        _log_event("info", context, op_id, "Flush Excel sin cola. Ruta=%s", _EXCEL_QUEUE)

        return {"procesados": 0, "pendientes": 0}



    lines = _EXCEL_QUEUE.read_text(encoding="utf-8").splitlines()

    if not lines:
        _log_event("info", context, op_id, "Flush Excel sin lineas. Ruta=%s", _EXCEL_QUEUE)

        return {"procesados": 0, "pendientes": 0}



    pendientes = []

    procesados = 0

    _log_event("info", context, op_id, "Flush Excel iniciado. Total=%s", len(lines))



    for idx, line in enumerate(lines):

        try:

            record = json.loads(line)

        except json.JSONDecodeError:

            _log_event("warning", context, op_id, "Linea invalida en cola. Ruta=%s", _EXCEL_QUEUE)

            pendientes.append(line)

            continue



        action = record.get("action", "append")

        ods = record.get("ods", {})

        original = record.get("original") or ods

        try:

            if action == "append":

                append_row(ods)

            elif action == "update":

                update_row(original, ods)

            elif action == "delete":

                delete_row(original)

            elif action == "rebuild":

                rebuild_excel_from_supabase_query(create_backup=False)

            else:

                raise RuntimeError(f"accion desconocida: {action}")

            procesados += 1

            _log_event("info", context, op_id, "Flush Excel procesado. Accion=%s", action)

        except PermissionError:

            _log_event("warning", context, op_id, "Flush Excel detenido por archivo abierto.")

            pendientes.append(line)

            pendientes.extend(lines[idx + 1 :])

            break

        except (RuntimeError, OSError, ValueError, TypeError):

            _log_event("exception", context, op_id, "Flush Excel fallo. Accion=%s", action)

            pendientes.append(line)



    _EXCEL_QUEUE.write_text("\n".join(pendientes) + ("\n" if pendientes else ""), encoding="utf-8")

    _log_event("info", context, op_id, "Flush Excel terminado. Procesados=%s Pendientes=%s", procesados, len(pendientes))

    return {"procesados": procesados, "pendientes": len(pendientes)}



