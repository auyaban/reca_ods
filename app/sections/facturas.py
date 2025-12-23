from collections import defaultdict
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

from app.supabase_client import get_supabase_client
from app.paths import app_data_dir, resource_path
from app.storage import ensure_appdata_files

router = APIRouter(prefix="/wizard/facturas", tags=["wizard"])

_OUTPUT_DIR = app_data_dir() / "facturas" / "generadas"


class FacturaItem(BaseModel):
    codigo_servicio: str
    referencia_servicio: str
    descripcion_servicio: str
    valor_base: float
    cantidad: float
    total: float


class CrearFacturaRequest(BaseModel):
    tipo: str  # "clausulada" o "no_clausulada"
    items: list[FacturaItem]
    nombre_archivo: str | None = None


class GenerarFacturaRequest(BaseModel):
    mes: int
    año: int
    tipo: str  # "clausulada" o "no_clausulada"
    nombre_archivo: str | None = None


class PreviewFacturaRequest(BaseModel):
    mes: int
    año: int
    tipo: str


def _get_template_path(tipo: str) -> Path:
    clean = tipo.strip().lower()
    ensure_appdata_files()
    if clean == "clausulada":
        path = app_data_dir() / "facturas" / "clausulada.xlsx"
        return path if path.exists() else resource_path("facturas/clausulada.xlsx")
    if clean in {"no_clausulada", "no-clausulada", "no clausulada"}:
        path = app_data_dir() / "facturas" / "no_clausulada.xlsx"
        return path if path.exists() else resource_path("facturas/no_clausulada.xlsx")
    raise HTTPException(status_code=422, detail="tipo de factura invalido")


def _find_first_empty_row(ws, start_row: int) -> int:
    for row_idx in range(start_row, ws.max_row + 2):
        row = ws[row_idx]
        if all(cell.value in (None, "") for cell in row[:6]):
            return row_idx
    return ws.max_row + 1


def _calcular_items(mes: int, año: int, tipo: str) -> list[FacturaItem]:
    tipo_clean = tipo.strip().lower()
    clausulada = tipo_clean == "clausulada"

    client = get_supabase_client()
    try:
        ods = (
            client.table("ods")
            .select(
                "codigo_servicio,referencia_servicio,descripcion_servicio,horas_interprete,orden_clausulada,mes_servicio,año_servicio"
            )
            .eq("mes_servicio", mes)
            .eq("año_servicio", año)
            .execute()
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Supabase error: {exc}") from exc

    registros = ods.data or []
    if not registros:
        raise HTTPException(status_code=404, detail="No se encontraron servicios para el periodo")

    filtrados = []
    for row in registros:
        orden = str(row.get("orden_clausulada", "")).strip().lower()
        es_clausulada = orden.startswith("s") or orden == "true"
        if es_clausulada == clausulada:
            filtrados.append(row)

    if not filtrados:
        raise HTTPException(status_code=404, detail="No hay servicios para el tipo solicitado")

    codigos = sorted(
        {str(row.get("codigo_servicio", "")).strip() for row in filtrados if row.get("codigo_servicio")}
    )
    if not codigos:
        raise HTTPException(status_code=422, detail="No hay codigos en el periodo")

    try:
        tarifas = (
            client.table("tarifas")
            .select("codigo_servicio,valor_base")
            .in_("codigo_servicio", codigos)
            .execute()
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Supabase error: {exc}") from exc

    valor_por_codigo = {
        str(item.get("codigo_servicio")): float(item.get("valor_base") or 0)
        for item in (tarifas.data or [])
    }

    agrupados = defaultdict(lambda: {"cantidad": 0.0, "horas": 0.0})
    meta = {}
    for row in filtrados:
        codigo = str(row.get("codigo_servicio", "")).strip()
        referencia = str(row.get("referencia_servicio", "")).strip()
        descripcion = str(row.get("descripcion_servicio", "")).strip()
        key = (codigo, referencia, descripcion)
        meta[key] = {"codigo": codigo, "referencia": referencia, "descripcion": descripcion}
        horas = row.get("horas_interprete") or 0
        try:
            horas_val = float(horas or 0)
        except (TypeError, ValueError):
            horas_val = 0.0
        if horas_val > 0:
            agrupados[key]["horas"] += horas_val
        else:
            agrupados[key]["cantidad"] += 1

    items: list[FacturaItem] = []
    for key, agg in agrupados.items():
        codigo = meta[key]["codigo"]
        referencia = meta[key]["referencia"]
        descripcion = meta[key]["descripcion"]
        valor_base = valor_por_codigo.get(codigo, 0.0)
        cantidad = agg["horas"] if agg["horas"] > 0 else agg["cantidad"]
        total = valor_base * cantidad
        items.append(
            FacturaItem(
                codigo_servicio=codigo,
                referencia_servicio=referencia,
                descripcion_servicio=descripcion,
                valor_base=valor_base,
                cantidad=cantidad,
                total=total,
            )
        )

    items.sort(key=lambda item: item.codigo_servicio)
    return items


@router.post("/crear")
def crear_factura(payload: CrearFacturaRequest) -> dict:
    if not payload.items:
        raise HTTPException(status_code=422, detail="Debes enviar items para la factura")

    template_path = _get_template_path(payload.tipo)
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="No se encontro la plantilla de factura")

    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl no esta instalado") from exc

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    header_row = 9
    start_row = _find_first_empty_row(ws, header_row)

    for idx, item in enumerate(payload.items, start=1):
        row_idx = start_row + (idx - 1)
        ws.cell(row=row_idx, column=1, value=item.codigo_servicio)  # Cod.
        ws.cell(row=row_idx, column=2, value=item.referencia_servicio)  # No
        ws.cell(row=row_idx, column=3, value=item.descripcion_servicio)
        ws.cell(row=row_idx, column=4, value=item.valor_base)
        ws.cell(row=row_idx, column=5, value=item.cantidad)
        ws.cell(row=row_idx, column=6, value=item.total)

    total_sum = sum(item.total for item in payload.items)
    ws["F45"] = total_sum
    ws["F46"] = round(total_sum * 0.19, 2)
    ws["F47"] = round(total_sum + ws["F46"].value, 2)

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if payload.nombre_archivo:
        filename = payload.nombre_archivo
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"factura_{payload.tipo}_{timestamp}.xlsx"
    output_path = _OUTPUT_DIR / filename

    wb.save(output_path)
    return {"data": {"archivo": str(output_path)}}


@router.post("/preview")
def preview_factura(payload: PreviewFacturaRequest) -> dict:
    if payload.mes < 1 or payload.mes > 12:
        raise HTTPException(status_code=422, detail="mes invalido")

    tipo = payload.tipo.strip().lower()
    if tipo not in {"clausulada", "no_clausulada", "no clausulada", "no-clausulada"}:
        raise HTTPException(status_code=422, detail="tipo de factura invalido")

    items = _calcular_items(payload.mes, payload.año, payload.tipo)
    total_sum = sum(item.total for item in items)
    iva = round(total_sum * 0.19, 2)
    total = round(total_sum + iva, 2)
    return {
        "data": {
            "items": [item.model_dump() for item in items],
            "total_base": total_sum,
            "iva": iva,
            "total": total,
        }
    }


@router.post("/generar")
def generar_factura(payload: GenerarFacturaRequest) -> dict:
    if payload.mes < 1 or payload.mes > 12:
        raise HTTPException(status_code=422, detail="mes invalido")

    tipo = payload.tipo.strip().lower()
    if tipo not in {"clausulada", "no_clausulada", "no clausulada", "no-clausulada"}:
        raise HTTPException(status_code=422, detail="tipo de factura invalido")

    items = _calcular_items(payload.mes, payload.año, payload.tipo)
    crear_payload = CrearFacturaRequest(tipo=payload.tipo, items=items, nombre_archivo=payload.nombre_archivo)
    return crear_factura(crear_payload)
