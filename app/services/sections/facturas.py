from datetime import datetime
from pathlib import Path

from pydantic import BaseModel, Field

from app.factura_calc import calcular_items
from app.factura_models import FacturaItem
from app.paths import app_data_dir, resource_path
from app.services.errors import ServiceError
from app.storage import ensure_appdata_files

_OUTPUT_DIR = app_data_dir() / "facturas" / "generadas"


class CrearFacturaRequest(BaseModel):
    tipo: str  # "clausulada" o "no_clausulada"
    items: list[FacturaItem]
    nombre_archivo: str | None = None


class GenerarFacturaRequest(BaseModel):
    mes: int
    aヵo: int = Field(alias="ano")
    tipo: str  # "clausulada" o "no_clausulada"
    nombre_archivo: str | None = None


class PreviewFacturaRequest(BaseModel):
    mes: int
    aヵo: int = Field(alias="ano")
    tipo: str


def _get_template_path(tipo: str) -> Path:
    clean = tipo.strip().lower()
    ensure_appdata_files()
    if clean == "clausulada":
        path = app_data_dir() / "facturas" / "clausulada.xlsx"
        return path if path.exists() else resource_path("facturas/clausulada.xlsx")
    if clean in {"no_clausulada", "no-clausulada", "no clausulada"}:
        path = app_data_dir() / "facturas" / "no_clausulada.xlsx"
        return path if path.exists() else resource_path("facturas/no_clausulada.xlsx")
    raise ServiceError("tipo de factura invalido", status_code=422)


def _find_first_empty_row(ws, start_row: int) -> int:
    for row_idx in range(start_row, ws.max_row + 2):
        row = ws[row_idx]
        if all(cell.value in (None, "") for cell in row[:6]):
            return row_idx
    return ws.max_row + 1


def crear_factura(payload: CrearFacturaRequest) -> dict:
    if not payload.items:
        raise ServiceError("Debes enviar items para la factura", status_code=422)

    template_path = _get_template_path(payload.tipo)
    if not template_path.exists():
        raise ServiceError("No se encontro la plantilla de factura", status_code=404)

    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise ServiceError("openpyxl no esta instalado", status_code=500) from exc

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    header_row = 9
    start_row = _find_first_empty_row(ws, header_row)

    for idx, item in enumerate(payload.items, start=1):
        row_idx = start_row + (idx - 1)
        ws.cell(row=row_idx, column=1, value=item.codigo_servicio)  # Cod.
        ws.cell(row=row_idx, column=2, value=item.referencia_servicio)  # No
        ws.cell(row=row_idx, column=3, value=item.descripcion_servicio)
        ws.cell(row=row_idx, column=4, value=item.valor_base)
        ws.cell(row=row_idx, column=5, value=item.cantidad)
        ws.cell(row=row_idx, column=6, value=item.total)

    total_sum = sum(item.total for item in payload.items)
    ws["F45"] = total_sum
    ws["F46"] = round(total_sum * 0.19, 2)
    ws["F47"] = round(total_sum + ws["F46"].value, 2)

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if payload.nombre_archivo:
        filename = payload.nombre_archivo
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"factura_{payload.tipo}_{timestamp}.xlsx"
    output_path = _OUTPUT_DIR / filename

    wb.save(output_path)
    return {"data": {"archivo": str(output_path)}}


def preview_factura(payload: PreviewFacturaRequest) -> dict:
    if payload.mes < 1 or payload.mes > 12:
        raise ServiceError("mes invalido", status_code=422)

    tipo = payload.tipo.strip().lower()
    if tipo not in {"clausulada", "no_clausulada", "no clausulada", "no-clausulada"}:
        raise ServiceError("tipo de factura invalido", status_code=422)

    items = calcular_items(payload.mes, payload.aヵo, payload.tipo)
    try:
        from app.excel_sync import update_factura_sheet

        update_factura_sheet(payload.mes, payload.aヵo, payload.tipo)
    except Exception:
        pass
    total_sum = sum(item.total for item in items)
    iva = round(total_sum * 0.19, 2)
    total = round(total_sum + iva, 2)
    return {
        "data": {
            "items": [item.model_dump() for item in items],
            "total_base": total_sum,
            "iva": iva,
            "total": total,
        }
    }


def generar_factura(payload: GenerarFacturaRequest) -> dict:
    if payload.mes < 1 or payload.mes > 12:
        raise ServiceError("mes invalido", status_code=422)

    tipo = payload.tipo.strip().lower()
    if tipo not in {"clausulada", "no_clausulada", "no clausulada", "no-clausulada"}:
        raise ServiceError("tipo de factura invalido", status_code=422)

    items = calcular_items(payload.mes, payload.aヵo, payload.tipo)
    crear_payload = CrearFacturaRequest(tipo=payload.tipo, items=items, nombre_archivo=payload.nombre_archivo)
    return crear_factura(crear_payload)


def debug_actualizar_factura(mes: int, ano: int, tipo: str) -> dict:
    try:
        from app.excel_sync import update_factura_sheet

        update_factura_sheet(mes, ano, tipo)
    except Exception as exc:
        raise ServiceError(f"Error al actualizar factura: {exc}", status_code=500) from exc
    return {"data": "ok"}
