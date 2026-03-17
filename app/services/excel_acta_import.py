from __future__ import annotations

import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.google_sheets_client import (
    download_drive_file,
    export_spreadsheet_to_excel,
    extract_drive_file_id,
    get_drive_file_metadata,
)
from app.utils.text import normalize_text

MAX_SCAN_ROWS = 260
MAX_SCAN_COLS = 70
_GOOGLE_SPREADSHEET_MIME = "application/vnd.google-apps.spreadsheet"
_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_PDF_EXTENSIONS = {".pdf"}
_PDF_BLOCK_RE = re.compile(
    r"(?ms)(?:^|\n)\s*(?P<idx>[1-9])\s+(?P<body>.*?)(?=(?:\n\s*[1-9]\s+[A-ZÃÃ‰ÃÃ“ÃšÃ‘])|\Z)"
)


def _clean_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_cedula(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    return digits


def _clean_nit(value: Any) -> str:
    raw = _clean_text(value)
    if not raw:
        return ""
    match = re.search(r"\d{6,12}(?:-\d)?", raw)
    return match.group(0) if match else raw


def _clean_name(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .:-")


def _looks_like_person_name_for_company(value: Any) -> bool:
    text = _clean_name(value)
    if not text or not _is_person_candidate(text):
        return False
    norm = normalize_text(text)
    company_markers = (
        "sas",
        "s a s",
        "sa",
        "ltda",
        "ips",
        "eps",
        "sucursal",
        "grupo",
        "group",
        "consult",
        "seguros",
        "rehabilit",
        "colombia",
        "partners",
        "soluciones",
        "solutions",
        "fundacion",
        "universidad",
        "clinica",
        "hospital",
        "colegio",
        "empresa",
    )
    if any(marker in norm for marker in company_markers):
        return False
    tokens = [token for token in re.split(r"\s+", text) if token]
    return 2 <= len(tokens) <= 4 and all(any(ch.isalpha() for ch in token) for token in tokens)


def _company_from_email_domain(text: str) -> str:
    email_match = re.search(
        r"(?i)[\w.+-]+@(?P<domain>[a-z0-9.-]+?\.(?:com\.co|edu\.co|org\.co|net\.co|com|org|net|co))",
        normalize_text(text or ""),
    )
    if not email_match:
        return ""
    domain = email_match.group("domain").split(".", 1)[0]
    pieces = re.split(r"[-_.]+", domain)
    if len(pieces) == 1:
        collapsed = pieces[0]
        for marker in (
            "rehabilitacion",
            "consulting",
            "seguros",
            "colombia",
            "partners",
            "solutions",
            "soluciones",
            "salud",
            "industrial",
            "tecnologia",
            "tecnologias",
            "servicios",
            "logistica",
        ):
            collapsed = re.sub(marker, f" {marker}", collapsed, flags=re.IGNORECASE)
        pieces = [part for part in collapsed.split() if part]
    company = " ".join(piece.upper() for piece in pieces if piece)
    return _clean_name(company)


def _to_iso_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _clean_text(value)
    if not text:
        return ""

    text = text.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _sheet_matrix(path: str) -> list[tuple[str, list[list[Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency
        raise RuntimeError("No se pudo importar openpyxl.") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        result: list[tuple[str, list[list[Any]]]] = []
        for ws in wb.worksheets:
            rows: list[list[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, max_col=MAX_SCAN_COLS, values_only=True):
                rows.append(list(row))
            result.append((ws.title, rows))
        return result
    finally:
        wb.close()


def _is_likely_label(text: str) -> bool:
    if not text:
        return False
    if len(text) > 70:
        return False
    return text.endswith(":") or any(token in text for token in ("nit", "fecha", "modalidad", "profesional"))


def _first_neighbor_value(rows: list[list[Any]], r: int, c: int) -> Any:
    # Right side first.
    for dc in range(1, 16):
        cc = c + dc
        if cc >= len(rows[r]):
            break
        value = rows[r][cc]
        if _clean_text(value):
            if len(_clean_text(value)) > 60:
                continue
            norm = normalize_text(value)
            if not _is_likely_label(norm):
                return value

    # Next rows under same/adjacent columns
    for dr in range(1, 3):
        rr = r + dr
        if rr >= len(rows):
            break
        for dc in (0, 1):
            cc = c + dc
            if cc >= len(rows[rr]):
                continue
            value = rows[rr][cc]
            if _clean_text(value):
                norm = normalize_text(value)
                if not _is_likely_label(norm):
                    return value
    return None


def _find_labeled_value(rows: list[list[Any]], label_tokens: tuple[str, ...], starts_with: bool = False) -> Any:
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            norm = normalize_text(value)
            if not norm:
                continue
            if len(norm) > 55:
                continue
            matched = any(norm.startswith(token) for token in label_tokens) if starts_with else any(
                token in norm for token in label_tokens
            )
            if matched:
                neighbor = _first_neighbor_value(rows, r, c)
                if _clean_text(neighbor):
                    return neighbor
    return None


def _is_person_candidate(value: Any) -> bool:
    text = _clean_text(value)
    if not text or len(text) < 3:
        return False
    norm = normalize_text(text)
    if re.search(r"(https?://|www\.|@[a-z0-9._-]+|\.com\b|\.org\b|\.net\b|\.co\b)", norm):
        return False
    banned = (
        "codigo",
        "tema",
        "version",
        "correo",
        "telefono",
        "nit",
        "empresa",
        "fecha",
        "modalidad",
        "objetivo",
        "cargo",
        "sede",
        "asesor",
        "direccion",
        "ciudad",
    )
    if any(token in norm for token in banned):
        return False
    return any(ch.isalpha() for ch in text)


def _extract_profesional(rows: list[list[Any]]) -> str:
    labels = (
        "profesional asignado reca",
        "profesional asignado",
    )
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            norm = normalize_text(value)
            if not norm or len(norm) > 60:
                continue
            if not any(norm == token or norm.startswith(token + ":") for token in labels):
                continue
            candidate = _first_neighbor_value(rows, r, c)
            if _is_person_candidate(candidate):
                return _clean_text(candidate)
    return ""


def _extract_asistentes_candidates(rows: list[list[Any]]) -> list[str]:
    asist_row = -1
    for r, row in enumerate(rows):
        for value in row:
            norm = normalize_text(value)
            if not norm:
                continue
            if re.search(r"\b\d+\s*\.\s*asistentes\b", norm) or norm == "asistentes" or norm.endswith(" asistentes"):
                asist_row = r
                break
        if asist_row >= 0:
            break

    if asist_row < 0:
        return []

    candidates: list[str] = []
    for rr in range(asist_row + 1, min(asist_row + 30, len(rows))):
        row = rows[rr]
        if not row:
            continue
        row_norm = [normalize_text(cell) for cell in row]
        has_nombre_label = any("nombre completo" in cell for cell in row_norm if cell)
        if not has_nombre_label:
            continue

        for raw in row:
            text = _clean_text(raw)
            norm = normalize_text(raw)
            if "nombre completo" in norm and ":" in text:
                inline = text.split(":", 1)[1].strip()
                if _is_person_candidate(inline):
                    candidates.append(inline)
                    break

        for raw in row:
            value = _clean_text(raw)
            norm = normalize_text(raw)
            if not value:
                continue
            if "nombre completo" in norm:
                continue
            if "cargo" in norm or "firma" in norm:
                continue
            if _is_person_candidate(value) and value not in candidates:
                candidates.append(value)
                break

    return candidates


def _extract_profesional_from_asistentes(rows: list[list[Any]]) -> str:
    candidates = _extract_asistentes_candidates(rows)
    return candidates[0] if candidates else ""


def _extract_nit(rows: list[list[Any]]) -> str:
    nit_labels = ("numero de nit", "nit empresa", "razon social / nit", "nit:")
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            raw = _clean_text(value)
            norm = normalize_text(value)
            if not raw:
                continue
            if len(norm) > 55:
                continue
            if not any(token in norm for token in nit_labels):
                continue
            own = _clean_nit(raw)
            if re.fullmatch(r"\d{6,12}(?:-\d)?", own):
                return own
            for dc in range(1, 9):
                cc = c + dc
                if cc >= len(row):
                    break
                candidate = re.sub(r"[.\s]", "", _clean_nit(row[cc]))
                if re.fullmatch(r"\d{6,12}(?:-\d)?", candidate):
                    return candidate
            for dr in range(1, 4):
                rr = r + dr
                if rr >= len(rows):
                    break
                candidate = re.sub(r"[.\s]", "", _clean_nit(rows[rr][c] if c < len(rows[rr]) else ""))
                if re.fullmatch(r"\d{6,12}(?:-\d)?", candidate):
                    return candidate
    return ""


def _extract_participants(rows: list[list[Any]]) -> list[dict[str, str]]:
    participants: list[dict[str, str]] = []
    name_headers = (
        "nombre vinculado",
        "nombre completo",
        "nombres y apellidos",
        "nombre del vinculado",
        "nombre usuario",
        "nombre participante",
        "nombre oferente",
    )
    ced_headers = ("cedula", "c.c", "cc", "documento")

    for idx, row in enumerate(rows):
        normalized = [normalize_text(cell) for cell in row]
        name_col = next((i for i, cell in enumerate(normalized) if any(token in cell for token in name_headers)), None)
        if name_col is None:
            continue
        ced_col = next((i for i, cell in enumerate(normalized) if any(token in cell for token in ced_headers)), None)
        if ced_col is None:
            continue
        discapacidad_col = next((i for i, cell in enumerate(normalized) if "discapacidad" in cell), None)
        genero_col = next((i for i, cell in enumerate(normalized) if "genero" in cell or "sexo" in cell), None)

        empty_streak = 0
        for j in range(idx + 1, min(idx + 120, len(rows))):
            current = rows[j]
            ced_raw = current[ced_col] if ced_col is not None and ced_col < len(current) else ""
            ced = _clean_cedula(ced_raw)

            if not ced:
                empty_streak += 1
                if empty_streak >= 4:
                    break
                continue
            empty_streak = 0

            if len(ced) < 5:
                continue

            name_raw = current[name_col] if name_col < len(current) else ""
            discapacidad_raw = current[discapacidad_col] if discapacidad_col is not None and discapacidad_col < len(current) else ""
            genero_raw = current[genero_col] if genero_col is not None and genero_col < len(current) else ""
            participants.append(
                {
                    "nombre_usuario": _clean_name(name_raw),
                    "cedula_usuario": ced,
                    "discapacidad_usuario": _clean_text(discapacidad_raw),
                    "genero_usuario": _clean_text(genero_raw),
                }
            )
    return participants


def _dedupe_participants(participants: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for item in participants:
        ced = _clean_cedula(item.get("cedula_usuario", ""))
        if not ced:
            continue
        key = ced.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(
            {
                "nombre_usuario": _clean_name(item.get("nombre_usuario", "")),
                "cedula_usuario": ced,
                "discapacidad_usuario": _clean_text(item.get("discapacidad_usuario", "")),
                "genero_usuario": _clean_text(item.get("genero_usuario", "")),
            }
        )
    return unique


def _extract_pdf_text_pages(path: str) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - runtime dependency
        raise RuntimeError("No se pudo importar pypdf.") from exc

    reader = PdfReader(path)
    pages: list[str] = []
    for page in reader.pages:
        raw_text = page.extract_text() or ""
        lines = [_clean_text(line) for line in raw_text.splitlines()]
        page_text = "\n".join(line for line in lines if line)
        pages.append(page_text)
    return pages


def _extract_pdf_value(text: str, pattern: str) -> str:
    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return ""
    return _clean_text(match.group(1))


def _extract_pdf_nits(text: str) -> list[str]:
    labeled_matches = re.findall(
        r"(?i)(?:numero de nit|n[uú]mero de nit|nit empresa|razon social / nit|nit)\s*:\s*([0-9.\- ]+)",
        str(text or ""),
    )
    seen: set[str] = set()
    result: list[str] = []
    for nit in labeled_matches:
        clean = _clean_nit(nit)
        if not clean or clean in seen:
            continue
        seen.add(clean)
        result.append(clean)
    if result:
        return result

    for nit in re.findall(r"\d{6,12}(?:-\d)?", str(text or "")):
        digits = re.sub(r"\D", "", nit)
        if len(digits) == 10 and digits.startswith("3"):
            continue
        clean = _clean_nit(nit)
        if not clean or clean in seen:
            continue
        seen.add(clean)
        result.append(clean)
    return result


def _parse_duration_hours(raw_value: str) -> float | None:
    text = normalize_text(raw_value or "")
    if not text:
        return None
    match = re.search(r"(\d+(?:[.,]\d+)?)", text)
    if not match:
        return None
    try:
        amount = float(match.group(1).replace(",", "."))
    except ValueError:
        return None
    if "min" in text:
        return round(amount / 60.0, 2)
    if "hora" in text:
        return round(amount, 2)
    return round(amount, 2)


def _extract_pdf_general_fields(first_page: str) -> tuple[str, str, str]:
    header_text = re.sub(
        r"(?i)(?<!\n)(fecha de la visita:|modalidad:|nombre de la empresa:|ciudad/municipio:|direcci[oó]n de la empresa:|n[uú]mero de nit:|correo electr[oó]nico:|tel[eé]fonos:|contacto de la empresa:|empresa afiliada a caja(?:de)? compensaci[oó]n:|sede compensar:|asesor:|profesional asignado\s*reca:)",
        r"\n\1",
        first_page,
    )
    empresa = _extract_pdf_value(
        header_text,
        r"nombre de la empresa:\s*([^\n]+?)(?:\s*(?:ciudad/municipio:|direcci[oó]n de la empresa:|n[uú]mero de nit:|correo electr[oó]nico:|tel[eé]fonos:|contacto de la empresa:|empresa afiliada a caja(?:de)? compensaci[oó]n:|$))",
    )
    fecha_servicio = _to_iso_date(
        _extract_pdf_value(header_text, r"fecha de la visita:\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})")
    )
    modalidad = _extract_pdf_value(
        header_text,
        r"modalidad:\s*([^\n]+?)(?:\s*(?:nombre de la empresa:|ciudad/municipio:|direcci[oó]n de la empresa:|n[uú]mero de nit:|correo electr[oó]nico:|tel[eé]fonos:|$))",
    ).rstrip(".")

    if normalize_text(empresa).startswith(("nombre de la empresa", "direccion de la empresa", "dirección de la empresa")):
        empresa = ""

    if empresa and fecha_servicio and modalidad:
        return empresa, fecha_servicio, modalidad

    lines = [_clean_text(line) for line in first_page.splitlines() if _clean_text(line)]
    for idx, line in enumerate(lines):
        normalized = normalize_text(line)
        if "modalidad:" not in normalized:
            continue

        if not fecha_servicio:
            date_match = re.search(
                r"(?P<fecha>[0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})\s+modalidad:",
                line,
                re.IGNORECASE,
            )
            if date_match:
                fecha_servicio = _to_iso_date(date_match.group("fecha"))

        if not modalidad:
            modalidad_match = re.search(r"modalidad:\s*(?P<modalidad>.+)$", line, re.IGNORECASE)
            if modalidad_match:
                modalidad = _clean_text(modalidad_match.group("modalidad")).rstrip(".")

        if not empresa and idx + 1 < len(lines):
            next_line = lines[idx + 1]
            if "ciudad/municipio:" in normalize_text(next_line):
                empresa = _clean_text(re.split(r"(?i)ciudad/municipio:", next_line, maxsplit=1)[0])

    if not empresa or normalize_text(empresa).startswith(
        ("nombre de la empresa", "direccion de la empresa", "dirección de la empresa")
    ):
        empresa = _company_from_email_domain(header_text)

    if not empresa:
        for idx, line in enumerate(lines):
            if normalize_text(line).startswith("fecha de la visita:") and idx > 0:
                previous_line = lines[idx - 1]
                if "ciudad/municipio:" in normalize_text(previous_line):
                    empresa = _clean_text(re.split(r"(?i)ciudad/municipio:", previous_line, maxsplit=1)[0])
                    break

    if not empresa or normalize_text(empresa).startswith(
        ("nombre de la empresa", "direccion de la empresa", "dirección de la empresa")
    ):
        empresa = _company_from_email_domain(header_text)

    return empresa, fecha_servicio, modalidad


def _extract_pdf_oferentes_section(text: str) -> str:
    match = re.search(
        r"(?is)(?<!\d)2\.\s*datos del oferente(?P<section>.*?)(?=(?<!\d)3\.\s*\S)",
        text,
    )
    if not match:
        return text
    return _clean_text(match.group("section"))


def _split_joined_cedula_percentage(raw_token: str) -> tuple[str, str]:
    match = re.search(r"(?P<digits>\d+)(?:(?P<sep>[.,])(?P<dec>\d{1,2}))?%?", raw_token)
    if not match:
        return "", ""

    digits = re.sub(r"\D", "", match.group("digits"))
    decimals = match.group("dec")
    separator = match.group("sep") or "."
    candidates: list[tuple[int, int, str, str]] = []
    for pct_len in (2, 1, 3):
        if len(digits) <= pct_len:
            continue
        cedula = digits[:-pct_len]
        percentage_int = digits[-pct_len:]
        if not (6 <= len(cedula) <= 10):
            continue
        percentage_value = int(percentage_int)
        if not (0 <= percentage_value <= 100):
            continue
        score = 0
        if pct_len == 2:
            score += 5
        if len(cedula) >= 8:
            score += 2
        if percentage_value > 0:
            score += 1
        percentage = f"{percentage_value}{separator}{decimals}%" if decimals is not None else f"{percentage_value}%"
        candidates.append((score, len(cedula), cedula, percentage))

    if not candidates:
        return "", ""

    _, _, cedula, percentage = max(candidates, key=lambda item: (item[0], item[1]))
    return cedula, percentage


def _split_joined_cedula_phone(raw_digits: str) -> tuple[str, str]:
    digits = re.sub(r"\D", "", raw_digits or "")
    if len(digits) < 17:
        return "", ""
    for cedula_len in (10, 9, 8, 7):
        if len(digits) <= cedula_len:
            continue
        cedula = digits[:cedula_len]
        phone = digits[cedula_len:]
        if len(phone) == 10 and phone.startswith("3"):
            return cedula, phone
    return "", ""


def _extract_pdf_participants(text: str) -> list[dict[str, str]]:
    participants: list[dict[str, str]] = []
    search_text = _extract_pdf_oferentes_section(text)

    for match in _PDF_BLOCK_RE.finditer(search_text):
        body = _clean_text(match.group("body"))
        if "discapacidad" not in normalize_text(body):
            continue
        first_line = body.split(" Agente de ", 1)[0]
        first_match = re.search(
            r"^(?P<nombre>.+?)(?P<token>\d{7,13}(?:[.,]\d{1,2})?%?)(?:\s*)Discapacidad\s+(?P<tail>.+)$",
            first_line,
            re.IGNORECASE,
        )
        if not first_match:
            continue

        nombre = _clean_name(first_match.group("nombre"))
        if not nombre or not _is_person_candidate(nombre):
            continue
        cedula, _pct = _split_joined_cedula_percentage(first_match.group("token"))
        if not cedula:
            continue

        tail = _clean_text(first_match.group("tail"))
        tail_match = re.search(
            r"^(?P<discapacidad>[^0-9]+?)(?P<telefono>\d[\d ]{6,15})(?P<resultado>Pendiente|Aprobado|No aprobado)",
            tail,
            re.IGNORECASE,
        )
        if not tail_match:
            continue
        discapacidad = _clean_text(tail_match.group("discapacidad")) if tail_match else ""
        participants.append(
            {
                "nombre_usuario": nombre,
                "cedula_usuario": cedula,
                "discapacidad_usuario": discapacidad,
                "genero_usuario": "",
            }
        )

    if participants:
        return _dedupe_participants(participants)

    inline_pattern = re.compile(
        r"(?is)(?<!\d)(?P<idx>[1-9])\s+"
        r"(?P<nombre>.+?)"
        r"(?P<token>\d{7,13}(?:[.,]\d{1,2})?%?)"
        r"(?:\s*)Discapacidad\s+"
        r"(?P<tail>.*?)(?=(?:(?<!\d)[1-9]\s+\S)|(?:(?<!\d)[3-9]\.\s*\S)|\Z)"
    )
    for match in inline_pattern.finditer(search_text):
        nombre = _clean_name(match.group("nombre"))
        if not nombre or not _is_person_candidate(nombre):
            continue
        cedula, _pct = _split_joined_cedula_percentage(match.group("token"))
        if not cedula:
            continue

        tail = _clean_text(match.group("tail"))
        tail_match = re.search(
            r"^(?P<discapacidad>[^0-9]+?)(?P<telefono>\d[\d ]{6,15})(?P<resultado>Pendiente|Aprobado|No aprobado)",
            tail,
            re.IGNORECASE,
        )
        if not tail_match:
            continue
        discapacidad = _clean_text(tail_match.group("discapacidad")) if tail_match else ""
        participants.append(
            {
                "nombre_usuario": nombre,
                "cedula_usuario": cedula,
                "discapacidad_usuario": discapacidad,
                "genero_usuario": "",
            }
        )
    if participants:
        return _dedupe_participants(participants)

    contract_pattern = re.compile(
        r"(?ims)(?:^|\n)\s*(?P<idx>[1-9])\s+"
        r"(?P<nombre>.+?)"
        r"(?P<token>\d{7,13}(?:[.,]\d{1,2})?%?)\s*"
        r"Discapacidad\s+"
        r"(?P<discapacidad>[^0-9]+?)"
        r"(?P<telefono>\d[\d ]{6,15})"
        r"(?=\s*(?:Masculino|Femenino|Otro)\b)"
    )
    for match in contract_pattern.finditer(search_text):
        nombre = _clean_name(match.group("nombre"))
        if not nombre or not _is_person_candidate(nombre):
            continue
        cedula, _pct = _split_joined_cedula_percentage(match.group("token"))
        if not cedula:
            continue
        participants.append(
            {
                "nombre_usuario": nombre,
                "cedula_usuario": cedula,
                "discapacidad_usuario": _clean_text(match.group("discapacidad")),
                "genero_usuario": "",
            }
        )
    if participants:
        return _dedupe_participants(participants)

    # Fallback for PDFs where the participant row is extracted outside the
    # expected section but still preserves name, cÃ©dula, discapacidad, phone,
    # and result on the same line.
    line_pattern = re.compile(
        r"(?i)(?P<nombre>[A-ZÃÃ‰ÃÃ“ÃšÃ‘][A-ZÃÃ‰ÃÃ“ÃšÃ‘a-zÃ¡Ã©Ã­Ã³ÃºÃ¼ÃœÃ±' .-]{8,}?)\s+"
        r"(?P<cedula>\d{6,12})\s+"
        r"Discapacidad\s+"
        r"(?P<discapacidad>.+?)\s+"
        r"(?P<telefono>\d[\d ]{6,15})\s+"
        r"(?P<resultado>Pendiente|Aprobado|No aprobado)\b"
    )
    for raw_line in str(text or "").splitlines():
        line = _clean_text(raw_line)
        if "discapacidad" not in normalize_text(line):
            continue
        match = line_pattern.search(line)
        if not match:
            continue
        nombre = _clean_name(match.group("nombre"))
        cedula = _clean_cedula(match.group("cedula"))
        if not nombre or not cedula or not _is_person_candidate(nombre):
            continue
        participants.append(
            {
                "nombre_usuario": nombre,
                "cedula_usuario": cedula,
                "discapacidad_usuario": _clean_text(match.group("discapacidad")),
                "genero_usuario": "",
            }
        )
    if participants:
        return _dedupe_participants(participants)

    follow_up_pattern = re.compile(
        r"(?is)persona que atiende la\s*visita.*?"
        r"(?P<nombre>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{8,}?)"
        r"(?P<cedula_phone>\d{17,22})"
        r"(?P<email>[\w.+-]+@[\w.-]+)"
        r"(?P<contacto>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{4,})"
        r"(?:Hermana|Hermano|Madre|Padre|Esposa|Esposo|Pareja|Amiga|Amigo|Tia|Tio|Prima|Primo)"
        r"\s+\d{7,12}\s+"
        r"(?P<cargo>.+?)\s+"
        r"Si\s+No aplica\.\s+Discapacidad\s+(?P<discapacidad>.+?)"
        r"(?=\s+\d{1,2}/\d{1,2}/\d{4}\b)",
    )
    follow_up_patterns = [
        follow_up_pattern,
        re.compile(
            r"(?is)(?P<nombre>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{8,}?)"
            r"(?P<cedula_phone>\d{17,22})"
            r"(?P<email>[\w.+-]+@[\w.-]+)"
            r"(?P<contacto>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{4,}?)"
            r"(?:Hermana|Hermano|Madre|Padre|Esposa|Esposo|Pareja|Amiga|Amigo|Tia|Tio|Prima|Primo)"
            r"\s+\d{7,12}\s+"
            r"(?P<cargo>.+?)\s+"
            r"Si\s+No aplica\.\s+Discapacidad\s+(?P<discapacidad>.+?)"
            r"(?=\s+(?:Seguimiento\s*[1-9]:|\d{1,2}/\d{1,2}/\d{4}\b)|$)",
        ),
        re.compile(
            r"(?is)(?P<nombre>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{8,}?)"
            r"(?P<cedula_phone>\d{17,22})"
            r"(?P<email>[\w.+-]+@[\w.-]+)"
            r"(?P<contacto>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúüÜñ' .-]{4,}?)"
            r"(?:Hermana|Hermano|Madre|Padre|Esposa|Esposo|Pareja|Amiga|Amigo|Tia|Tio|Prima|Primo)"
            r"\s+\d{7,12}\s+"
            r"(?P<cargo>.+?)\s+"
            r"Si\s+(?:(?P<porcentaje>\d{1,3}(?:[.,]\d{1,2})?)|No refiere|No aplica\.)\s+Discapacidad\s+(?P<discapacidad>.+?)"
            r"(?=\s+(?:Contrato de trabajo|Seguimiento\s*[1-9]:|\d{1,2}/\d{1,2}/\d{4}\b)|$)",
        ),
    ]
    for follow_up_pattern in follow_up_patterns:
        follow_up_match = follow_up_pattern.search(str(text or ""))
        if not follow_up_match:
            continue
        nombre = _clean_name(follow_up_match.group("nombre"))
        cedula, _telefono = _split_joined_cedula_phone(follow_up_match.group("cedula_phone"))
        discapacidad = _clean_text(follow_up_match.group("discapacidad"))
        if nombre and cedula and _is_person_candidate(nombre):
            participants.append(
                {
                    "nombre_usuario": nombre,
                    "cedula_usuario": cedula,
                    "discapacidad_usuario": discapacidad,
                    "genero_usuario": "",
                }
            )
            break
    return _dedupe_participants(participants)


def _extract_pdf_follow_up_number(text: str) -> str:
    last_number = ""
    for match in re.finditer(
        r"(?i)seguimiento\s*(?P<number>[1-9])\s*:\s*(?P<date>[0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})",
        str(text or ""),
    ):
        if _to_iso_date(match.group("date")):
            last_number = match.group("number")
    return last_number


def _extract_pdf_vacancy_fields(text: str) -> tuple[str, int]:
    cargo = _extract_pdf_value(
        text,
        r"(?is)nombre de la vacante:\s*(.+?)(?:n[uú]mero de vacantes:|nivel del cargo:|g[eé]nero:|edad:|modalidad de trabajo:|lugar de trabajo:|$)",
    )
    total_vacantes = 0
    vacantes_raw = _extract_pdf_value(
        text,
        r"(?is)n[uú]mero de vacantes:\s*(.+?)(?:nivel del cargo:|g[eé]nero:|edad:|modalidad de trabajo:|lugar de trabajo:|salario asignado:|$)",
    )
    vacantes_match = re.search(r"\d+", vacantes_raw)
    if vacantes_match:
        total_vacantes = int(vacantes_match.group(0))
    return _clean_text(cargo), total_vacantes


def _extract_pdf_selection_cargo(text: str) -> str:
    section_match = re.search(
        r"(?is)2\.\s*datos del oferente(?P<section>.*?)(?:(?<!\d)3\.\s*\S|4\.\s*caracterizaci[oó]n del oferente|$)",
        str(text or ""),
    )
    section = section_match.group("section") if section_match else str(text or "")
    compact = re.sub(r"\s+", " ", section)
    header_match = re.search(
        r"(?is)cargo\s*contacto de emergencia\s*parentesco\s*tel[eé]fono(?:\s*fecha de nacimiento\s*edad)?\s*(?P<tail>.+?)(?:(?:[¿?]pendiente otros oferentes|lugar firma de contrato|fecha firma de contrato|3\.)|$)",
        compact,
    )
    if header_match:
        tail = _clean_text(header_match.group("tail"))
        tail = re.sub(r"(?<=[a-záéíóúüñ])(?=[A-ZÁÉÍÓÚÑ])", " ", tail)
        contact_match = re.search(
            r"(?s)^(?P<cargo>.+?)\s+"
            r"(?P<contacto>[A-ZÁÉÍÓÚÑ][a-záéíóúüñ]+(?:\s+[A-Za-zÁÉÍÓÚÑáéíóúüñ]+){0,4})\s+"
            r"(?P<parentesco>Madre|Padre|Hermana|Hermano|Pareja|Esposa|Esposo|Amiga|Amigo|Mamá|Papa|Papá|Tia|Tío|Tio|Abuela|Abuelo)\b",
            tail,
        )
        cargo = _clean_text(contact_match.group("cargo") if contact_match else tail)
        if cargo:
            return cargo
    cargo = _extract_pdf_value(
        section,
        r"(?is)cargo\s+(?P<value>.+?)(?:contacto de emergencia|parentesco|tel[eé]fono|fecha de nacimiento|edad|[¿?]pendiente otros oferentes|lugar firma de contrato|fecha firma de contrato|$)",
    )
    if normalize_text(cargo).startswith("contacto de emergencia"):
        return ""
    return _clean_text(cargo)


def _extract_name_from_follow_up_filename(path: Path) -> str:
    stem = path.stem
    suffix_match = re.search(r"\s*-\s*\d{2}_[A-Za-z]{3,4}_\d{4}$", stem)
    if not suffix_match:
        return ""
    prefix = stem[: suffix_match.start()]
    if " - " not in prefix:
        return ""
    name = prefix.rsplit(" - ", 1)[-1]
    name = re.sub(r"^\(\d+\)\s*", "", name).strip()
    return _clean_name(name)


def _extract_follow_up_participant_from_filename(text: str, path: Path) -> list[dict[str, str]]:
    if "seguimiento al proceso de inclusion laboral" not in normalize_text(text):
        return []
    cedula_match = re.search(r"(?P<cedula_phone>\d{17,22})", str(text or ""), re.IGNORECASE)
    discapacidad_match = re.search(
        r"(?is)Si\s+(?:\d{1,3}(?:[.,]\d{1,2})?%?|No refiere|No aplica\.)\s+Discapacidad\s+(?P<discapacidad>.+?)"
        r"(?=\s+(?:Contrato de trabajo|Seguimiento\s*[1-9]:|\d{1,2}/\d{1,2}/\d{4}\b)|$)",
        str(text or ""),
    )
    if not cedula_match or not discapacidad_match:
        return []
    nombre = _extract_name_from_follow_up_filename(path)
    cedula, _telefono = _split_joined_cedula_phone(cedula_match.group("cedula_phone"))
    discapacidad = _clean_text(discapacidad_match.group("discapacidad"))
    if not nombre or not cedula:
        return []
    return [
        {
            "nombre_usuario": nombre,
            "cedula_usuario": cedula,
            "discapacidad_usuario": discapacidad,
            "genero_usuario": "",
        }
    ]


def _extract_pdf_asistentes_candidates(text: str) -> list[str]:
    def _collect(source_text: str) -> list[str]:
        found: list[str] = []
        normalized = re.sub(r"(?i)(?<!\n)(nombre completo:)", r"\n\1", source_text)
        for raw_line in normalized.splitlines():
            line = raw_line.strip()
            if not line or "nombre completo:" not in line.lower():
                continue
            explicit_name = re.search(
                r"(?i)nombre completo:\s*(?P<nombre>[A-ZÃÃ‰ÃÃ“ÃšÃ‘][A-Za-zÃÃ‰ÃÃ“ÃšÃ‘Ã¡Ã©Ã­Ã³ÃºÃ¼ÃœÃ±' -]+?)"
                r"(?:\s+(?:cargo:|profesional\b|lider\b|coordinacion\b|psicolog[aÃ¡]\b)|$)",
                line,
            )
            if explicit_name:
                candidate = _clean_name(explicit_name.group("nombre"))
                if candidate and _is_person_candidate(candidate) and candidate not in found:
                    found.append(candidate)
                continue
            for chunk in re.split(r"(?i)nombre completo:\s*", line):
                candidate = re.split(r"(?i)\bcargo:\s*", chunk, maxsplit=1)[0]
                candidate = _clean_name(candidate)
                if candidate and _is_person_candidate(candidate) and candidate not in found:
                    found.append(candidate)
        return found

    start_match = re.search(r"\b\d+\.\s*asistentes\b", text, re.IGNORECASE)
    asistentes_text = text[start_match.start() :] if start_match else text
    candidates = _collect(asistentes_text)
    if candidates or not start_match:
        return candidates
    return _collect(text)


def _extract_interpreter_names(text: str) -> list[str]:
    names: list[str] = []
    pattern = re.compile(
        r"(?is)nombre\s+int\S?rprete\s*(?:no\s*\d+)?\s*:\s*(?P<nombre>.+?)(?:hora\s+inicial:|hora\s+final:|total\s+tiempo:|\n|$)"
    )
    for match in pattern.finditer(str(text or "")):
        candidate = _clean_name(match.group("nombre"))
        if candidate and _is_person_candidate(candidate) and candidate not in names:
            names.append(candidate)
    return names


def _extract_interpreter_participants(text: str) -> tuple[list[dict[str, str]], str]:
    normalized_text = str(text or "")
    section_match = re.search(
        r"(?is)2\.\s*datos de los oferentes/ vinculados(?P<section>.*?)(?:nombre\s+int\S?rprete|3\.)",
        normalized_text,
    )
    section = _clean_text(section_match.group("section")) if section_match else _clean_text(normalized_text)
    process_name = ""
    participants: list[dict[str, str]] = []
    pattern = re.compile(
        r"(?P<idx>\d+)\s+(?P<nombre>.+?)\s+(?P<cedula>\d{6,12})\s+(?P<proceso>.+?)(?=(?:\s+\d+\s+[A-Z????????????])|$)",
        re.IGNORECASE,
    )
    for match in pattern.finditer(section):
        nombre = _clean_name(match.group("nombre"))
        cedula = _clean_cedula(match.group("cedula"))
        proceso = _clean_text(match.group("proceso"))
        if not nombre or not cedula:
            continue
        if not process_name and proceso:
            process_name = proceso
        participants.append(
            {
                "nombre_usuario": nombre,
                "cedula_usuario": cedula,
                "discapacidad_usuario": "",
                "genero_usuario": "",
            }
        )
    return _dedupe_participants(participants), process_name


def _parse_interpreter_pdf(first_page: str, full_text: str, path: Path) -> dict:
    asistentes_candidates = _extract_pdf_asistentes_candidates(full_text)
    fecha_servicio = _to_iso_date(
        _extract_pdf_value(
            full_text,
            r"1\.\s*datos de la empresa\s*fecha:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        )
    )
    empresa = _extract_pdf_value(
        full_text,
        r"nombre de la empresa:\s*(.+?)(?:direcci[oó]n:|contacto en la empresa:|modalidad servicio:|2\.)",
    )
    modalidad_match = re.search(
        r"(?is)modalidad servicio:\s*(?:(?:int\S?rprete|nombre int\S?rprete)\s*:\s*.+?\s+)?"
        r"(?P<modalidad>virtual|presencial|h[ií]brida?)\b",
        full_text,
    )
    modalidad = _clean_text(modalidad_match.group("modalidad")) if modalidad_match else ""
    profesional_reca = _extract_pdf_value(
        full_text,
        r"profesional reca:\s*(.+?)(?:\bvirtual\b|\bpresencial\b|\bh[ií]brida?\b|2\.|3\.)",
    )
    interpreter_names = _extract_interpreter_names(full_text)
    participants, process_name = _extract_interpreter_participants(full_text)
    sumatoria_raw = _extract_pdf_value(
        full_text,
        r"(?is)sumatoria horas int\S?rpretes:\s*(.+?)(?:observaciones:|3\.)",
    )
    total_time_raw = _extract_pdf_value(
        full_text,
        r"(?is)total tiempo:\s*(.+?)(?:si el servicio fue realizado en sabana|sumatoria horas int\S?rpretes:)",
    )
    total_time_hours = _parse_duration_hours(total_time_raw)
    sumatoria_hours = _parse_duration_hours(sumatoria_raw)
    horas = sumatoria_hours if sumatoria_hours is not None else total_time_hours
    nit = _clean_nit(_extract_pdf_value(first_page, r"n\S?mero de nit:\s*([0-9.\- ]+)"))
    nits = [nit] if nit else []

    warnings: list[str] = []
    if not empresa:
        warnings.append("No se detecto nombre de empresa en el PDF.")
    if not fecha_servicio:
        warnings.append("No se detecto fecha de servicio en formato valido.")
    if not participants:
        warnings.append("No se detectaron oferentes en el PDF.")
    if horas is None:
        warnings.append("No se detecto total de horas interprete en el PDF.")
    if not interpreter_names and not asistentes_candidates:
        warnings.append("No se detectaron interpretes ni asistentes en el PDF.")

    return {
        "file_path": str(path),
        "nit_empresa": nit,
        "nits_empresas": nits,
        "nombre_empresa": empresa,
        "fecha_servicio": fecha_servicio,
        "nombre_profesional": interpreter_names[0] if interpreter_names else (asistentes_candidates[0] if asistentes_candidates else (profesional_reca or "")),
        "interpretes": interpreter_names,
        "candidatos_profesional": interpreter_names or asistentes_candidates,
        "asistentes": asistentes_candidates,
        "modalidad_servicio": modalidad,
        "participantes": participants,
        "interpreter_process_name": process_name,
        "interpreter_total_time_raw": total_time_raw,
        "sumatoria_horas_interpretes_raw": sumatoria_raw,
        "total_horas_interprete": total_time_hours if total_time_hours is not None else "",
        "sumatoria_horas_interpretes": sumatoria_hours if sumatoria_hours is not None else "",
        "is_fallido": "fallido" in normalize_text(full_text),
        "warnings": warnings,
    }

def parse_acta_excel(file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise RuntimeError(f"No existe el archivo: {file_path}")

    sheets = _sheet_matrix(str(path))

    nit = ""
    empresa = ""
    fecha_servicio = ""
    profesional = ""
    modalidad = ""
    participants: list[dict[str, str]] = []
    candidatos_profesional: list[str] = []

    for _, rows in sheets:
        if not nit:
            nit = _extract_nit(rows)
        if not empresa:
            empresa_value = _find_labeled_value(
                rows,
                ("nombre de la empresa", "razon social"),
                starts_with=True,
            )
            empresa = _clean_text(empresa_value)
        if not fecha_servicio:
            fecha_value = _find_labeled_value(
                rows,
                ("fecha de la visita", "fecha servicio", "fecha de firma de contrato", "fecha firma de contrato"),
                starts_with=True,
            )
            fecha_servicio = _to_iso_date(fecha_value)
        sheet_candidates = _extract_asistentes_candidates(rows)
        for candidate in sheet_candidates:
            if candidate not in candidatos_profesional:
                candidatos_profesional.append(candidate)
        if not profesional:
            profesional = sheet_candidates[0] if sheet_candidates else _extract_profesional(rows)
        if not modalidad:
            modalidad_value = _find_labeled_value(
                rows,
                ("modalidad:", "modalidad"),
                starts_with=True,
            )
            modalidad = _clean_text(modalidad_value)

        participants.extend(_extract_participants(rows))

    participants = _dedupe_participants(participants)

    warnings: list[str] = []
    if not nit:
        warnings.append("No se detecto NIT en el archivo.")
    if not fecha_servicio:
        warnings.append("No se detecto fecha de servicio en formato valido.")

    return {
        "file_path": str(path),
        "nit_empresa": nit,
        "nombre_empresa": empresa,
        "fecha_servicio": fecha_servicio,
        "nombre_profesional": profesional,
        "candidatos_profesional": candidatos_profesional,
        "modalidad_servicio": modalidad,
        "participantes": participants,
        "warnings": warnings,
    }


def parse_acta_pdf(file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise RuntimeError(f"No existe el archivo: {file_path}")

    pages = _extract_pdf_text_pages(str(path))
    if not pages:
        raise RuntimeError("El PDF no contiene paginas legibles.")

    full_text = "\n".join(page for page in pages if page)
    first_page = pages[0] if pages else ""
    normalized_first_page = normalize_text(first_page)
    if "interprete" in normalize_text(first_page) and "sumatoria horas interpretes" in normalize_text(full_text):
        return _parse_interpreter_pdf(first_page, full_text, path)

    nit = _clean_nit(_extract_pdf_value(first_page, r"n(?:u|ú|Ãº)mero de nit:\s*([0-9.\- ]+)"))
    empresa, fecha_servicio, modalidad = _extract_pdf_general_fields(first_page)
    asistentes_candidates = _extract_pdf_asistentes_candidates(full_text)
    profesional_reca = _extract_pdf_value(
        first_page,
        r"profesional asignado\s*reca:\s*(.*?)(?:modalidad:|\n|se informa|$)",
    )
    asesor = _extract_pdf_value(first_page, r"asesor:\s*(.+?)(?:sede compensar:|correo electr[oÃ³]nico:|$)")
    if normalize_text(profesional_reca).startswith("modalidad"):
        profesional_reca = ""
    profesional = asistentes_candidates[0] if asistentes_candidates else (profesional_reca or asesor)
    participants = _extract_pdf_participants(full_text)
    if not participants:
        participants = _extract_follow_up_participant_from_filename(full_text, path)
    nits = _extract_pdf_nits(full_text)
    numero_seguimiento = _extract_pdf_follow_up_number(full_text)
    cargo_objetivo, total_vacantes = _extract_pdf_vacancy_fields(full_text)
    if not cargo_objetivo and "proceso de seleccion incluyente" in normalized_first_page:
        cargo_objetivo = _extract_pdf_selection_cargo(full_text)
    if not nit and nits:
        nit = nits[0]

    warnings: list[str] = []
    if not nit:
        warnings.append("No se detecto NIT en el PDF.")
    if not empresa:
        warnings.append("No se detecto nombre de empresa en el PDF.")
    if not fecha_servicio:
        warnings.append("No se detecto fecha de servicio en formato valido.")
    if (
        not participants
        and "evaluacion de accesibilidad" not in normalized_first_page
        and "revision de las condiciones de la vacante" not in normalized_first_page
    ):
        warnings.append("No se detectaron oferentes en el PDF.")

    return {
        "file_path": str(path),
        "nit_empresa": nit,
        "nits_empresas": nits,
        "nombre_empresa": empresa,
        "fecha_servicio": fecha_servicio,
        "nombre_profesional": profesional,
        "candidatos_profesional": asistentes_candidates or ([profesional_reca] if profesional_reca else ([asesor] if asesor else [])),
        "modalidad_servicio": modalidad,
        "cargo_objetivo": cargo_objetivo,
        "total_vacantes": total_vacantes,
        "numero_seguimiento": numero_seguimiento,
        "participantes": participants,
        "warnings": warnings,
    }


def _is_google_spreadsheet_reference(source: str) -> bool:
    text = str(source or "").strip().lower()
    return "/spreadsheets/d/" in text


def _is_google_drive_reference(source: str) -> bool:
    text = str(source or "").strip().lower()
    return "drive.google.com" in text and ("/file/d/" in text or "id=" in text)


def _parse_local_source(path: Path) -> dict:
    suffix = path.suffix.lower()
    if suffix in _EXCEL_EXTENSIONS:
        return parse_acta_excel(str(path))
    if suffix in _PDF_EXTENSIONS:
        return parse_acta_pdf(str(path))
    raise RuntimeError("El archivo local no es un Excel compatible (.xlsx/.xlsm) ni un PDF compatible.")


def _normalize_source_type_label(source_type: str, parsed: dict, source_text: str) -> dict:
    parsed["file_path"] = source_text
    parsed["source_type"] = source_type
    return parsed


def _parse_google_remote_source(source_text: str) -> dict:
    file_id = extract_drive_file_id(source_text)
    metadata = get_drive_file_metadata(file_id)
    mime_type = str(metadata.get("mimeType") or "").strip().lower()
    name = str(metadata.get("name") or "acta").strip()
    suffix = Path(name).suffix.lower()
    temp_path: Path | None = None
    try:
        if mime_type == _GOOGLE_SPREADSHEET_MIME:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                temp_path = Path(tmp.name)
            export_spreadsheet_to_excel(file_id, temp_path)
            parsed = parse_acta_excel(str(temp_path))
            return _normalize_source_type_label("google_sheets", parsed, source_text)

        if suffix not in (_EXCEL_EXTENSIONS | _PDF_EXTENSIONS):
            raise RuntimeError(
                "El archivo de Drive no es un Google Sheet, un Excel compatible (.xlsx/.xlsm) ni un PDF compatible."
            )

        with tempfile.NamedTemporaryFile(suffix=suffix or ".xlsx", delete=False) as tmp:
            temp_path = Path(tmp.name)
        download_drive_file(file_id, temp_path)
        parser = parse_acta_pdf if suffix in _PDF_EXTENSIONS or mime_type == "application/pdf" else parse_acta_excel
        parsed = parser(str(temp_path))
        return _normalize_source_type_label("google_drive_file", parsed, source_text)
    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def parse_acta_source(source: str) -> dict:
    source_text = str(source or "").strip()
    if not source_text:
        raise RuntimeError("Debe indicar la ruta o URL del acta.")

    local_path = Path(source_text).expanduser()
    if local_path.exists():
        return _parse_local_source(local_path)

    if _is_google_spreadsheet_reference(source_text) or _is_google_drive_reference(source_text):
        return _parse_google_remote_source(source_text)

    raise RuntimeError(
        "No se pudo resolver el acta. Usa un archivo Excel/PDF local o una URL valida de Google Drive/Sheets."
    )


