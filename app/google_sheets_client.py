from __future__ import annotations

import os
import re
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any

from app.config import get_settings
from app.utils.cache import ttl_bucket

_CLIENT_CACHE_TTL_SECONDS = 300
_SHEETS_DRIVE_SCOPES = (
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
)
_GMAIL_READONLY_SCOPES = (
    "https://www.googleapis.com/auth/gmail.readonly",
)
_SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
_GOOGLE_DOC_FILE_ID_RE = re.compile(r"https?://docs\.google\.com/(?:spreadsheets|document|presentation|forms)/d/([a-zA-Z0-9-_]+)", re.IGNORECASE)
_DRIVE_FILE_ID_RE = re.compile(r"/file/d/([a-zA-Z0-9-_]+)|[?&]id=([a-zA-Z0-9-_]+)")
_GOOGLE_FILE_FIELDS = "files(id,name,mimeType,parents,driveId),nextPageToken"
_GOOGLE_SPREADSHEET_MIME = "application/vnd.google-apps.spreadsheet"


def extract_spreadsheet_id(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise RuntimeError("Debe indicar spreadsheet_id o URL de Google Sheets.")

    match = _SPREADSHEET_ID_RE.search(text)
    if match:
        return match.group(1)
    return text


def extract_drive_file_id(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise RuntimeError("Debe indicar file_id o URL de Google Drive.")

    match = _DRIVE_FILE_ID_RE.search(text)
    if match:
        return match.group(1) or match.group(2) or ""
    match = _GOOGLE_DOC_FILE_ID_RE.search(text)
    if match:
        return match.group(1)
    return text


def normalize_google_file_open_url(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise RuntimeError("Debe indicar una URL o file_id de Google Drive.")

    file_id = extract_drive_file_id(text)
    metadata = get_drive_file_metadata(file_id)
    mime_type = str(metadata.get("mimeType") or "").strip().lower()
    if mime_type == _GOOGLE_SPREADSHEET_MIME:
        return f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
    return f"https://drive.google.com/file/d/{file_id}/view"


def _credentials_path() -> Path:
    settings = get_settings()
    credential_path = settings.google_service_account_file
    if not credential_path:
        raise RuntimeError("Falta GOOGLE_SERVICE_ACCOUNT_FILE en la configuracion.")

    path = Path(os.path.expandvars(credential_path)).expanduser()
    if not path.exists():
        raise RuntimeError(f"No existe el archivo de service account: {path}")
    return path


@lru_cache
def _get_google_credentials_cached(
    credential_path: str,
    scopes_key: tuple[str, ...],
    delegated_subject: str,
    _ttl_bucket: int,
):
    try:
        from google.oauth2.service_account import Credentials
    except ImportError as exc:
        raise RuntimeError(
            "Faltan dependencias de Google Sheets/Drive. Instala requirements.txt."
        ) from exc

    credentials = Credentials.from_service_account_file(credential_path, scopes=scopes_key)
    if delegated_subject:
        credentials = credentials.with_subject(delegated_subject)
    return credentials


@lru_cache
def _get_google_service_cached(
    api_name: str,
    api_version: str,
    credential_path: str,
    scopes_key: tuple[str, ...],
    delegated_subject: str,
    _ttl_bucket: int,
):
    try:
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Faltan dependencias de Google Sheets/Drive. Instala requirements.txt."
        ) from exc

    credentials = _get_google_credentials_cached(credential_path, scopes_key, delegated_subject, _ttl_bucket)
    return build(api_name, api_version, credentials=credentials, cache_discovery=False)


def clear_google_sheets_service_cache() -> None:
    _get_google_credentials_cached.cache_clear()
    _get_google_service_cached.cache_clear()


def _get_google_service(api_name: str, api_version: str, *, scopes: tuple[str, ...], delegated_subject: str = ""):
    path = _credentials_path()
    bucket = ttl_bucket(_CLIENT_CACHE_TTL_SECONDS)
    return _get_google_service_cached(api_name, api_version, str(path), tuple(scopes), delegated_subject.strip(), bucket)


def get_google_sheets_service():
    return _get_google_service("sheets", "v4", scopes=_SHEETS_DRIVE_SCOPES)


def get_google_drive_service():
    return _get_google_service("drive", "v3", scopes=_SHEETS_DRIVE_SCOPES)


def get_google_gmail_service(*, delegated_subject: str | None = None):
    settings = get_settings()
    subject = str(delegated_subject or settings.google_gmail_delegated_user or "").strip()
    if not subject:
        raise RuntimeError("Falta GOOGLE_GMAIL_DELEGATED_USER para acceder a Gmail con service account.")
    return _get_google_service("gmail", "v1", scopes=_GMAIL_READONLY_SCOPES, delegated_subject=subject)


def get_default_spreadsheet_id() -> str:
    settings = get_settings()
    spreadsheet_id = settings.google_sheets_default_spreadsheet_id
    if not spreadsheet_id:
        raise RuntimeError(
            "Falta GOOGLE_SHEETS_DEFAULT_SPREADSHEET_ID en la configuracion."
        )
    return extract_spreadsheet_id(spreadsheet_id)


def get_spreadsheet(spreadsheet_id_or_url: str, *, include_grid_data: bool = False) -> dict[str, Any]:
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    return (
        service.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, includeGridData=include_grid_data)
        .execute()
    )


def read_sheet_values(spreadsheet_id_or_url: str, range_name: str) -> list[list[Any]]:
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    response = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=range_name)
        .execute()
    )
    return list(response.get("values", []))


def write_sheet_values(
    spreadsheet_id_or_url: str,
    range_name: str,
    values: list[list[Any]],
    *,
    value_input_option: str = "USER_ENTERED",
) -> dict[str, Any]:
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    body = {"values": values}
    return (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption=value_input_option,
            body=body,
        )
        .execute()
    )


def clear_sheet_values(spreadsheet_id_or_url: str, range_name: str) -> dict[str, Any]:
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    return (
        service.spreadsheets()
        .values()
        .clear(spreadsheetId=spreadsheet_id, range=range_name, body={})
        .execute()
    )


def list_drive_files(
    *,
    folder_id: str,
    name: str | None = None,
    mime_type: str | None = None,
) -> list[dict[str, Any]]:
    drive = get_google_drive_service()
    query_parts = [f"'{folder_id}' in parents", "trashed = false"]
    if name:
        safe_name = str(name).replace("'", "\\'")
        query_parts.append(f"name = '{safe_name}'")
    if mime_type:
        safe_mime = str(mime_type).replace("'", "\\'")
        query_parts.append(f"mimeType = '{safe_mime}'")

    files: list[dict[str, Any]] = []
    page_token: str | None = None
    while True:
        response = (
            drive.files()
            .list(
                q=" and ".join(query_parts),
                fields=_GOOGLE_FILE_FIELDS,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                pageToken=page_token,
            )
            .execute()
        )
        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return files


def copy_drive_file(
    file_id: str,
    *,
    new_name: str,
    parent_folder_id: str,
) -> dict[str, Any]:
    drive = get_google_drive_service()
    body = {
        "name": new_name,
        "parents": [parent_folder_id],
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    return (
        drive.files()
        .copy(
            fileId=file_id,
            body=body,
            supportsAllDrives=True,
            fields="id,name,mimeType,parents,driveId",
        )
        .execute()
    )


def get_drive_file_metadata(file_id_or_url: str) -> dict[str, Any]:
    file_id = extract_drive_file_id(file_id_or_url)
    drive = get_google_drive_service()
    return (
        drive.files()
        .get(
            fileId=file_id,
            supportsAllDrives=True,
            fields="id,name,mimeType,parents,driveId,webViewLink",
        )
        .execute()
    )


def download_drive_file(file_id_or_url: str, destination: str | Path) -> Path:
    file_id = extract_drive_file_id(file_id_or_url)
    drive = get_google_drive_service()
    destination_path = Path(destination).expanduser()
    destination_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError as exc:
        raise RuntimeError(
            "Faltan dependencias de Google Sheets/Drive. Instala requirements.txt."
        ) from exc

    request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    buffer = BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _status, done = downloader.next_chunk()
    destination_path.write_bytes(buffer.getvalue())
    return destination_path


def export_spreadsheet_to_excel(spreadsheet_id_or_url: str, destination: str | Path) -> Path:
    spreadsheet = get_spreadsheet(spreadsheet_id_or_url, include_grid_data=False)
    destination_path = Path(destination).expanduser()
    destination_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl no esta instalado.") from exc

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {})
        title = str(props.get("title") or "Sheet")
        ws = wb.create_sheet(title=title[:31] or "Sheet")
        rows = read_sheet_values(spreadsheet["spreadsheetId"], f"'{title}'")
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        frozen_rows = int(props.get("gridProperties", {}).get("frozenRowCount", 0) or 0)
        frozen_cols = int(props.get("gridProperties", {}).get("frozenColumnCount", 0) or 0)
        if frozen_rows > 0 or frozen_cols > 0:
            ws.freeze_panes = ws.cell(row=frozen_rows + 1, column=frozen_cols + 1)

    if not wb.worksheets:
        wb.create_sheet("Sheet")
    wb.save(destination_path)
    return destination_path
